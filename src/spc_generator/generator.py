"""
 SPC GENERATOR
 ----------------------------------------------------------------------
 Logic:
  1. SCANS for tabs starting with "SPC_"
  2. EXTRACTS HEADER ROWS to find Part #, Batch, Date, etc.
  3. CALCULATES Analysis (Stats + Rules + Cpk)
  4. GENERATES PDF with SECTION BREAKS per Tab.
  
 Version: 4.8.2 (Dynamic Headers, Strict Indexing, Blank Nominal Support)
"""

# --- IMPORTS ---
import sys
import os
import glob
import time
import re
from datetime import datetime
from pathlib import Path

# --- CRITICAL FIX FOR THREADING ERROR ---
import matplotlib
matplotlib.use('Agg') 
import matplotlib.pyplot as plt

# Data Science
import pandas as pd
import numpy as np
from scipy.stats import norm

# Excel
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# PDF Reporting
from fpdf import FPDF

# TUI Imports
import questionary
from rich.console import Console
from rich.panel import Panel
from rich.progress import Progress, SpinnerColumn, TextColumn, BarColumn
from rich.tree import Tree
from rich import box

console = Console()

# --- CONSTANTS ---
TOOL_VERSION = "4.8.2"
INPUT_PREFIX = "SPC_"        
INSERT_START_ROW = 8   
HEADER_SEARCH_ROWS = 7 

# --- COLORS ---
COLOR_PURPLE_DARK = "7030A0"   
COLOR_PURPLE_LIGHT = "8064A2"  

# --- PDF ENGINE CLASS ---
class WhitePaperPDF(FPDF):
    def __init__(self, filename):
        super().__init__()
        self.report_filename = filename
        self.set_auto_page_break(auto=True, margin=15)

    def header(self):
        self.set_font('Arial', 'B', 14)
        self.set_text_color(50, 50, 50)
        self.cell(0, 10, 'Process Capability Engineering Report', 0, 1, 'L')
        self.set_draw_color(0, 0, 0)
        self.line(10, 20, 200, 20)
        self.ln(10)

    def footer(self):
        self.set_y(-15)
        self.set_font('Arial', 'I', 8)
        self.set_text_color(128, 128, 128)
        self.cell(0, 10, f'File: {self.report_filename} | Page {self.page_no()}', 0, 0, 'C')

    def chapter_title(self, title, subtitle=None):
        self.set_font('Arial', 'B', 12)
        self.set_fill_color(230, 240, 255) # Light Blue
        self.set_text_color(0, 0, 0)
        self.cell(0, 8, f"  {title}", 0, 1, 'L', 1)
        if subtitle:
            self.set_font('Arial', 'I', 10)
            self.cell(0, 6, f"  {subtitle}", 0, 1, 'L')
        self.ln(4)

    def add_section_header(self, sheet_name, metadata):
        """Creates a divider page for a new Tab Section"""
        self.add_page()
        self.set_font('Arial', 'B', 16)
        self.set_fill_color(112, 48, 160) # Purple
        self.set_text_color(255, 255, 255)
        self.cell(0, 12, f"  SECTION: {sheet_name}", 0, 1, 'L', 1)
        self.ln(5)
        
        # Metadata Table
        if metadata:
            self.set_text_color(0, 0, 0)
            self.set_font('Arial', 'B', 12)
            self.cell(0, 8, "Run Information:", 0, 1, 'L')
            self.ln(2)
            
            self.set_font('Arial', 'B', 10)
            self.set_fill_color(240, 240, 240)
            
            # Draw Table
            col_w_key = 50
            col_w_val = 130
            
            for key, val in metadata.items():
                self.set_font('Arial', 'B', 10)
                self.cell(col_w_key, 8, f"  {key}", 1, 0, 'L', 1)
                self.set_font('Arial', '', 10)
                self.cell(col_w_val, 8, f"  {val}", 1, 1, 'L')
            
            self.ln(10)

    def add_stat_table(self, metrics):
        self.set_font('Arial', 'B', 10)
        self.set_fill_color(240, 240, 240)
        self.set_text_color(0, 0, 0)
        
        col_w = 45
        self.cell(col_w, 7, "Metric", 1, 0, 'C', 1)
        self.cell(col_w, 7, "Value", 1, 1, 'C', 1)
        
        self.set_font('Arial', '', 10)
        for name, val in metrics:
            self.cell(col_w, 7, f"  {name}", 1)
            self.cell(col_w, 7, f"  {val}", 1)
            self.ln()
        self.ln(5)

    def body_text(self, txt):
        self.set_font('Times', '', 11)
        self.set_text_color(0, 0, 0)
        self.multi_cell(0, 5, txt)
        self.ln()

# --- UTILITIES ---
def get_unique_filepath(filepath):
    if not os.path.exists(filepath): return filepath
    folder, filename = os.path.split(filepath)
    name, ext = os.path.splitext(filename)
    counter = 1
    while os.path.exists(os.path.join(folder, f"{name}_{counter}{ext}")):
        counter += 1
    return os.path.join(folder, f"{name}_{counter}{ext}")

def wait_for_file_access(filepath):
    if not os.path.exists(filepath): return
    retries = 0
    while retries < 3: 
        try:
            with open(filepath, 'a'): break
        except IOError:
            console.print(f"[yellow]WAITING:[/yellow] File {os.path.basename(filepath)} is open. Close it to continue.")
            time.sleep(2)
            retries += 1
    else:
        raise PermissionError("File is currently open in another program (like Excel). Please close it and run again.")

def sanitize_filename(filename):
    return re.sub(r'[\\/*?:"<>|]', "", str(filename))

def extract_sheet_metadata(ws):
    metadata = {}
    target_keys = ["PART", "BATCH", "DATE", "NOTE", "OPERATOR", "MACHINE", "ORDER", "LOT"]
    for row in ws.iter_rows(min_row=1, max_row=HEADER_SEARCH_ROWS, min_col=1, max_col=10):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                val_str = str(cell.value).strip().rstrip(':')
                if any(k in val_str.upper() for k in target_keys):
                    next_col_idx = cell.column + 1
                    neighbor = ws.cell(row=cell.row, column=next_col_idx).value
                    if neighbor:
                        metadata[val_str] = str(neighbor)
    return metadata

def calculate_cpk(data, usl, lsl):
    if len(data) < 2: return 0, 0, 0, 0
    mean = np.mean(data)
    sigma = np.std(data, ddof=1)
    if sigma < 1e-9: sigma = 1e-9 
    cpu = (usl - mean) / (3 * sigma)
    cpl = (mean - lsl) / (3 * sigma)
    cpk = min(cpu, cpl)
    cp = (usl - lsl) / (6 * sigma)
    return cp, cpk, mean, sigma

# --- EXCEL FORMATTING UTILS ---
def style_header_cell(cell):
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color=COLOR_PURPLE_LIGHT, end_color=COLOR_PURPLE_LIGHT, fill_type="solid")
    cell.alignment = Alignment(horizontal="center")
    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

def style_data_cell(cell, is_nominal=False, is_numeric=False, align_left=False):
    cell.alignment = Alignment(horizontal='left' if align_left else 'center')
    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    if is_nominal: cell.font = Font(bold=True)
    if is_numeric: cell.number_format = '0.0000'

def shift_existing_images(ws, insertion_row, count):
    if not hasattr(ws, '_images'): return
    for img in ws._images:
        try:
            if hasattr(img.anchor, '_from'):
                current_row = img.anchor._from.row
                if current_row >= (insertion_row - 1):
                    img.anchor._from.row += count
                    if hasattr(img.anchor, 'to'):
                        img.anchor.to.row += count
        except Exception: pass

def write_rule_legend(ws, start_row):
    legend_data = [
        ("WECO Rule 1", "Any single point outside 3σ limit"),
        ("WECO Rule 2", "2 of 3 consecutive points > 2σ (same side)"),
        ("WECO Rule 3", "4 of 5 consecutive points > 1σ (same side)"),
        ("WECO Rule 4", "8 consecutive points on one side of Mean"),
        ("Trend", "6 consecutive points increasing or decreasing"),
        ("Mixture", "8 consecutive points > 1σ (avoiding center)"),
        ("Stratification", "15 consecutive points < 1σ (hugging center)"),
        ("Alternating", "14 consecutive points alternating up/down")
    ]
    
    # Shifted to Column N (14) to avoid the new Dev columns
    col_start = 14 
    ws.column_dimensions['N'].width = 15
    ws.column_dimensions['O'].width = 55
    
    head_r = ws.cell(start_row, col_start, "RULE REFERENCE")
    head_r.font = Font(bold=True, size=14, color="FFFFFF")
    head_r.fill = PatternFill(start_color=COLOR_PURPLE_DARK, end_color=COLOR_PURPLE_DARK, fill_type="solid")
    
    r = start_row + 1
    style_header_cell(ws.cell(r, col_start, "Rule / Pattern"))
    style_header_cell(ws.cell(r, col_start+1, "Description"))
    
    r += 1
    for name, desc in legend_data:
        c1 = ws.cell(r, col_start, name)
        c2 = ws.cell(r, col_start+1, desc)
        style_data_cell(c1, is_nominal=True)
        style_data_cell(c2, align_left=True)
        r += 1

# --- SPC LOGIC ENGINE ---
def check_spc_rules_full_scan(data, mean, std_dev):
    try:
        n = len(data)
        if n == 0: return "INSUFFICIENT DATA", False, []
        if n == 1: return f"LIMITED DATA (N=1)", False, [] 
        if std_dev == 0: return "NO VARIATION (σ=0)", False, []

        z = (data - mean) / std_dev
        
        r1_indices = np.where(np.abs(z) > 3)[0]
        if len(r1_indices) > 0: return f"WECO Rule 1 (Sample #{r1_indices[0]+1})", True, r1_indices

        if n >= 3:
            for i in range(n - 2):
                window = z[i:i+3]
                if np.sum(window > 2) >= 2 or np.sum(window < -2) >= 2:
                    return f"WECO Rule 2 (Samples {i+1}-{i+3})", True, range(i, i+3)

        if n >= 5:
            for i in range(n - 4):
                window = z[i:i+5]
                if np.sum(window > 1) >= 4 or np.sum(window < -1) >= 4:
                    return f"WECO Rule 3 (Samples {i+1}-{i+5})", True, range(i, i+5)

        if n >= 8:
            for i in range(n - 7):
                window = z[i:i+8]
                if np.all(window > 0) or np.all(window < 0):
                    return f"WECO Rule 4 (Samples {i+1}-{i+8})", True, range(i, i+8)

        if n >= 6:
            for i in range(n - 5):
                window = data[i:i+6]
                diffs = np.diff(window)
                if np.all(diffs > 0) or np.all(diffs < 0):
                    return f"Trend Detected (Samples {i+1}-{i+6})", True, range(i, i+6)

        if n < 8: return f"LIMITED DATA (N={n})", False, []
        return "STABLE", False, []

    except Exception:
        return "CALC ERROR", False, []

# --- PLOTTING 1: CONTROL CHARTS ---
def create_summary_image(data_array, title, output_folder, index, prefix, usl_val, lsl_val):
    fig = plt.figure(figsize=(10, 5), dpi=100)
    ax = fig.add_subplot(111)
    ax.axis('off')
    count = len(data_array)
    if count > 0:
        val_avg = np.mean(data_array)
        text_str = (
            f"DATA SUMMARY: {title}\n"
            f"Samples: {count}\n"
            f"Range: {lsl_val:.4f} - {usl_val:.4f}\n"
            f"Average: {val_avg:.4f}\n"
            f"(Insufficient Data for Chart)"
        )
    else:
        text_str = f"NO DATA FOUND"
    ax.text(0.1, 0.5, text_str, transform=ax.transAxes, fontsize=12, fontfamily='monospace')
    clean_name = sanitize_filename(title)
    save_path = os.path.join(output_folder, f"TEMP_SUM_{prefix}_{clean_name}_{index}.png")
    plt.savefig(save_path, bbox_inches='tight')
    plt.close(fig)
    return save_path

# --- PLOTTING 2: BELL CURVES ---
def create_bell_curve_plot(data, usl, lsl, nominal_disp, mean, sigma, feature_name, sheet_name, output_dir, prefix):
    if sigma <= 1e-9:
        plt.figure(figsize=(10, 6))
        plt.text(0.5, 0.5, "Cannot generate Bell Curve:\nInsufficient Variation or Data", 
                 horizontalalignment='center', verticalalignment='center', transform=plt.gca().transAxes, fontsize=14)
        plt.title(f'Capability Analysis: {feature_name}', fontsize=12, fontweight='bold')
        filename = f"TEMP_BELL_{prefix}_{sanitize_filename(feature_name)}.png"
        save_path = os.path.join(output_dir, filename)
        plt.savefig(save_path, bbox_inches='tight', dpi=100)
        plt.close()
        return save_path

    plt.figure(figsize=(10, 6))
    effective_sigma = max(sigma, (usl-lsl)/20) 
    spread = max((usl-lsl), (6*effective_sigma)) * 1.5
    
    x = np.linspace(mean - spread/2, mean + spread/2, 1000)
    y = norm.pdf(x, mean, sigma)
    
    plt.plot(x, y, color='blue', linewidth=2, label=f'Process (σ={sigma:.4f})')
    plt.fill_between(x, y, alpha=0.2, color='blue')
    
    plt.axvline(lsl, color='red', linestyle='--', linewidth=2, label=f'LSL ({lsl})')
    plt.axvline(usl, color='red', linestyle='--', linewidth=2, label=f'USL ({usl})')
    
    if isinstance(nominal_disp, (int, float)):
        plt.axvline(nominal_disp, color='black', linestyle=':', linewidth=1, label=f'Nominal ({nominal_disp})')
        
    plt.axvline(mean, color='green', linestyle='-', linewidth=1, label=f'Mean ({mean:.4f})')

    plt.title(f'Capability Analysis: {feature_name}', fontsize=12, fontweight='bold')
    plt.xlabel('Measured Value')
    plt.ylabel('Probability Density')
    plt.legend(loc='best')
    plt.grid(True, alpha=0.3)
    
    filename = f"TEMP_BELL_{prefix}_{sanitize_filename(feature_name)}.png"
    save_path = os.path.join(output_dir, filename)
    plt.savefig(save_path, bbox_inches='tight', dpi=100)
    plt.close()
    return save_path

# --- PROCESSOR ---
def process_single_file(filepath, output_dir):
    filename = os.path.basename(filepath)
    wait_for_file_access(filepath)
    
    date_str = datetime.now().strftime("%Y%m%d")
    suffix = filename[9:] if filename.startswith("SPC-DATA_") else filename
    target_filename = f"{date_str}_SPC-RESULTS_{suffix}"
    pdf_filename = f"{date_str}_REPORT_{suffix}.pdf"
    
    target_path = output_dir / target_filename
    output_path = get_unique_filepath(str(target_path))
    pdf_path = get_unique_filepath(str(output_dir / pdf_filename))
    
    tab_log = []

    try:
        wb_output = load_workbook(filepath)
    except Exception as e:
        return {"critical_error": f"Could not open Excel file: {str(e)}", "logs": []}

    xls = pd.ExcelFile(filepath)
    all_sheet_names = xls.sheet_names
    
    processed_tabs = []
    temp_files = [] 
    
    pdf = WhitePaperPDF(filename)
    pdf.add_page()
    pdf.chapter_title(f"Executive Summary")
    pdf.body_text(f"File: {filename}")
    pdf.body_text(f"Processed on: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    pdf.ln(5)
    
    pdf_summary_data = []

    for sheet_name in all_sheet_names:
        if not sheet_name.startswith(INPUT_PREFIX):
            tab_log.append({'name': sheet_name, 'status': 'SKIP', 'msg': f"Tab name must start with '{INPUT_PREFIX}'"})
            continue

        try:
            ws_meta = wb_output[sheet_name]
            metadata = extract_sheet_metadata(ws_meta)
            
            # --- DYNAMIC HEADER DETECTION ---
            df_temp = pd.read_excel(filepath, sheet_name=sheet_name, header=None)
            header_row_idx = 7 # Fallback default
            for i in range(min(20, len(df_temp))):
                first_cell = str(df_temp.iloc[i, 0]).strip().upper()
                if first_cell in ["FEATURE NAME", "FEATURE", "DIMENSION", "DIM"]:
                    header_row_idx = i
                    break
                    
            df_raw = pd.read_excel(filepath, sheet_name=sheet_name, header=header_row_idx)

            if df_raw.shape[1] > 0 and df_raw.columns[0] not in df_raw.index.names:
                df_raw.set_index(df_raw.columns[0], inplace=True)

            # --- STRICT LABEL IDENTIFICATION ---
            idx_nom, idx_usl, idx_lsl, idx_ut, idx_lt = None, None, None, None, None
            for i in df_raw.index:
                val = str(i).strip().upper()
                if not idx_nom and val in ["NOMINAL", "NOM", "TARGET"]: idx_nom = i
                elif not idx_usl and val in ["USL", "MAX", "UPPER LIMIT", "UPPER SPEC"]: idx_usl = i
                elif not idx_lsl and val in ["LSL", "MIN", "LOWER LIMIT", "LOWER SPEC"]: idx_lsl = i
                elif not idx_ut and ("UPPER" in val and "TOL" in val): idx_ut = i
                elif not idx_lt and ("LOWER" in val and "TOL" in val): idx_lt = i

            has_limits = (idx_usl is not None and idx_lsl is not None)
            has_tols = (idx_ut is not None and idx_lt is not None)

            if not idx_nom and not has_limits:
                tab_log.append({'name': sheet_name, 'status': 'SKIP', 'msg': "Missing 'Nominal' row AND missing 'USL/LSL' rows."})
                continue
            
            if not has_limits and not has_tols:
                 tab_log.append({'name': sheet_name, 'status': 'SKIP', 'msg': "Could not find limits. Need rows for 'USL/LSL' OR 'Upper/Lower Tolerance'"})
                 continue

            feature_cols = [c for c in df_raw.columns if "Unnamed" not in str(c)]
            metadata_keys = [x for x in [idx_nom, idx_usl, idx_lsl, idx_ut, idx_lt] if x is not None]
            
            features_data = []
            
            for col in feature_cols:
                try:
                    # 1. Extract Nominal
                    nom_val_float = None
                    if idx_nom is not None:
                        nom_raw = df_raw.loc[idx_nom, col]
                        if pd.notna(nom_raw):
                            nom_str = str(nom_raw).strip().upper()
                            if nom_str not in ["", "NAN", "NULL", "NONE", "N/A", "-"]:
                                try: nom_val_float = float(nom_raw)
                                except ValueError: pass
                    
                    has_nom = (nom_val_float is not None)

                    # 2. Extract Absolute Limits
                    usl_val, lsl_val = None, None
                    if has_limits:
                        raw_usl = df_raw.loc[idx_usl, col]
                        raw_lsl = df_raw.loc[idx_lsl, col]
                        if pd.notna(raw_usl) and pd.notna(raw_lsl):
                            u_str = str(raw_usl).strip().upper()
                            l_str = str(raw_lsl).strip().upper()
                            if u_str not in ["", "NAN", "-", "N/A"] and l_str not in ["", "NAN", "-", "N/A"]:
                                try:
                                    usl_val = float(raw_usl)
                                    lsl_val = float(raw_lsl)
                                except ValueError: pass

                    # 3. Extract Tolerances
                    if (usl_val is None or lsl_val is None) and has_tols and has_nom:
                        raw_ut = df_raw.loc[idx_ut, col]
                        raw_lt = df_raw.loc[idx_lt, col]
                        if pd.notna(raw_ut) and pd.notna(raw_lt):
                            ut_str = str(raw_ut).strip().upper()
                            lt_str = str(raw_lt).strip().upper()
                            if ut_str not in ["", "NAN", "-", "N/A"] and lt_str not in ["", "NAN", "-", "N/A"]:
                                try:
                                    usl_val = nom_val_float + abs(float(raw_ut))
                                    lsl_val = nom_val_float - abs(float(raw_lt))
                                except ValueError: pass

                    if usl_val is None or lsl_val is None:
                        continue 
                        
                    # 5. Missing Nominal Midpoint
                    if not has_nom:
                        nom_val_float = (usl_val + lsl_val) / 2.0

                    if usl_val == lsl_val:
                        usl_val += 0.0001
                        lsl_val -= 0.0001

                    # 6. Extract Samples Safely
                    sample_rows = df_raw[~df_raw.index.isin(metadata_keys)]
                    subset_indices = sample_rows.index.tolist()
                    subset_values = sample_rows[col].tolist()
                    
                    clean_samples = []
                    plot_split_locs = [] 
                    
                    for i, (label, val) in enumerate(zip(subset_indices, subset_values)):
                        if str(label).upper().strip() == "SPLIT" or str(val).upper().strip() == "SPLIT":
                            plot_split_locs.append(len(clean_samples))
                            continue
                        try:
                            num = float(val)
                            if not np.isnan(num): 
                                clean_samples.append(num)
                        except (ValueError, TypeError): pass
                            
                    cp, cpk, mean_val, sigma_val = calculate_cpk(np.array(clean_samples), usl_val, lsl_val)
                    
                    features_data.append({
                        'name': col, 
                        'nominal': nom_val_float, 
                        'nominal_disp': nom_val_float if has_nom else "N/A", 
                        'usl': usl_val, 'lsl': lsl_val,
                        'data': np.array(clean_samples),
                        'split_locs': plot_split_locs,
                        'cp': cp, 'cpk': cpk, 'mean': mean_val, 'sigma': sigma_val
                    })
                    
                    pdf_summary_data.append([sheet_name, col, f"{cpk:.2f}", "PASS" if cpk >= 1.33 else "FAIL"])

                except Exception as e:
                    tab_log.append({'name': sheet_name, 'status': 'ERR', 'msg': f"Col {col} crashed: {str(e)}"})
                    continue
            
            if not features_data: 
                tab_log.append({'name': sheet_name, 'status': 'SKIP', 'msg': "No valid columns found (Check USL/LSL/Nominal)"})
                continue

            # --- PART 3: EXCEL GENERATION ---
            ws = wb_output[sheet_name]
            processed_tabs.append(sheet_name)
            safe_name = sanitize_filename(sheet_name)[:20]

            ws['A1'] = "SPC ANALYSIS RESULTS"
            ws['A1'].fill = PatternFill(start_color=COLOR_PURPLE_DARK, end_color=COLOR_PURPLE_DARK, fill_type="solid")

            num_features = len(features_data)
            table_rows_needed = 2 + num_features + 1
            chart_rows_needed = num_features 
            total_insert_count = table_rows_needed + chart_rows_needed + 2

            shift_existing_images(ws, INSERT_START_ROW, total_insert_count)
            ws.insert_rows(INSERT_START_ROW, amount=total_insert_count)
            
            # Align legend with the start of the Analysis Summary table
            write_rule_legend(ws, INSERT_START_ROW) 

            # Updated Columns and Widths
            cols = ['A','B','C','D','E','F','G','H', 'I', 'J', 'K', 'L']
            widths = [25, 15, 15, 15, 15, 15, 30, 20, 15, 15, 15, 15]
            for l, w in zip(cols, widths): ws.column_dimensions[l].width = w
            
            r = INSERT_START_ROW
            ws.cell(r,1,"ANALYSIS SUMMARY").font = Font(bold=True, size=14, color="FFFFFF")
            ws.cell(r,1).fill = PatternFill(start_color=COLOR_PURPLE_DARK, end_color=COLOR_PURPLE_DARK, fill_type="solid")
            r += 1
            
            headers = ["Feature", "Nominal", "USL (Calc)", "LSL (Calc)", "Mean", "StdDev", "Pattern / Status", "OOT Points", "Dev. Tol (+)", "Dev. Tol (-)", "Dev. USL", "Dev. LSL"]
            for i, h in enumerate(headers, 1): style_header_cell(ws.cell(r, i, h))
            
            f_pass, font_p = PatternFill(start_color="C6EFCE", fill_type="solid"), Font(color="006100", bold=True)
            f_warn, font_w = PatternFill(start_color="FFEB9C", fill_type="solid"), Font(color="9C6500", bold=True)
            f_fail, font_f = PatternFill(start_color="FFC7CE", fill_type="solid"), Font(color="9C0006", bold=True)
            
            r += 1 
            for feat in features_data:
                data = feat['data']
                has_data = len(data) >= 1
                
                mean = np.mean(data) if has_data else 0
                std_dev = np.std(data, ddof=1) if has_data and len(data) > 1 else 0
                
                status, is_unstable, highlight_idx = check_spc_rules_full_scan(data, mean, std_dev)
                oot_count = np.sum(data > feat['usl']) + np.sum(data < feat['lsl']) if has_data else 0
                oot_disp = f"{oot_count} FAILED ({(oot_count/len(data))*100:.1f}%)" if oot_count > 0 else 0

                # New logic integrated here
                dev_tol_minus = ""
                dev_tol_plus = ""
                dev_usl = ""
                dev_lsl = ""

                if has_data:
                    data_min = np.min(data)
                    data_max = np.max(data)
                    
                    dev_usl = data_max
                    dev_lsl = data_min
                    
                    if feat['nominal_disp'] != "N/A":
                        if data_min < feat['lsl']: 
                            dev_tol_minus = feat['nominal'] - data_min
                        if data_max > feat['usl']: 
                            dev_tol_plus = data_max - feat['nominal']

                row_vals = [
                    str(feat['name']), feat['nominal_disp'], feat['usl'], feat['lsl'], 
                    mean, std_dev, status, oot_disp, dev_tol_plus, dev_tol_minus, dev_usl, dev_lsl
                ]
                
                for c_idx, val in enumerate(row_vals, 1):
                    cell = ws.cell(r, c_idx, val)
                    is_num_col = (has_data and c_idx in [2,3,4,5,6,9,10,11,12] and isinstance(val, (int, float)))
                    style_data_cell(cell, is_nominal=(c_idx==2), is_numeric=is_num_col)
                    
                    if c_idx == 5 and has_data:
                        tb = feat['usl'] - feat['lsl']
                        if mean > feat['usl'] or mean < feat['lsl']: cell.fill, cell.font = f_fail, font_f
                        elif mean > (feat['usl'] - 0.1*tb) or mean < (feat['lsl'] + 0.1*tb): cell.fill, cell.font = f_warn, font_w
                        else: cell.fill, cell.font = f_pass, font_p
                    if c_idx == 7 and has_data:
                        if is_unstable: cell.fill, cell.font = f_fail, font_f
                        elif "LIMITED" in status or "NO VAR" in status: cell.fill, cell.font = f_warn, font_w
                        else: cell.fill, cell.font = f_pass, font_p
                    if c_idx == 8 and has_data:
                        if oot_count > 0: cell.fill, cell.font = f_fail, font_f
                        else: cell.fill, cell.font = f_pass, font_p
                r += 1

            r += 1 
            img_start_row = r
            
            pdf.add_section_header(sheet_name, metadata)

            for i, feat in enumerate(features_data):
                data = feat['data']
                name = str(feat['name'])
                has_data = len(data) >= 1
                unique_prefix = f"{safe_name}_{i}"
                
                # --- NO DATA HANDLING ---
                if not has_data:
                    cell_name = ws.cell(img_start_row, 1, name)
                    cell_name.font, cell_name.alignment = Font(bold=True, size=12), Alignment(vertical='top')
                    
                    cell_msg = ws.cell(img_start_row, 2, "NO DATA FOUND")
                    cell_msg.font = Font(bold=True, italic=True, color="808080") # Grey, italic text
                    cell_msg.alignment = Alignment(vertical='top')
                    
                    ws.row_dimensions[img_start_row].height = 15 # Keep row height normal
                    img_start_row += 1
                    continue # Skip the rest of the image generation for this feature
                
                # --- PDF REPORT GENERATION (Only runs if has_data is True) ---
                pdf.add_page()
                pdf.chapter_title(f"Feature: {name}", subtitle=f"Sheet: {sheet_name}")
                
                bell_path = create_bell_curve_plot(data, feat['usl'], feat['lsl'], feat['nominal_disp'], 
                                          feat['mean'], feat['sigma'], name, sheet_name, output_dir, f"BELL_{unique_prefix}")
                temp_files.append(bell_path)
                
                pdf.image(bell_path, x=15, w=170)
                pdf.ln(5)
                
                status_text = "CAPABLE" if feat['cpk'] >= 1.33 else "NOT CAPABLE"
                conclusion = (
                    f"The process is statistically {status_text} (Cpk {feat['cpk']:.2f}). "
                    f"The mean is centered at {feat['mean']:.4f}, with a standard deviation of {feat['sigma']:.4f}. "
                )
                
                pdf.chapter_title("Statistical Interpretation")
                pdf.body_text(conclusion)
                pdf.ln(5)
                
                nom_str_pdf = f"{feat['nominal_disp']:.4f}" if isinstance(feat['nominal_disp'], (int, float)) else str(feat['nominal_disp'])
                
                metrics = [
                    ("Nominal", nom_str_pdf),
                    ("Tolerance", f"{feat['lsl']:.4f} to {feat['usl']:.4f}"),
                    ("Process Mean", f"{feat['mean']:.4f}"),
                    ("Sigma (Est)", f"{feat['sigma']:.5f}"),
                    ("Cp (Potential)", f"{feat['cp']:.2f}"),
                    ("Cpk (Actual)", f"{feat['cpk']:.2f}")
                ]
                pdf.add_stat_table(metrics)

                # --- EXCEL CHART GENERATION ---
                mean = np.mean(data)
                std_dev = np.std(data, ddof=1) if len(data) > 1 else 1e-9
                _, is_unstable, highlight_indices = check_spc_rules_full_scan(data, mean, std_dev)

                fig, ax = plt.subplots(figsize=(10, 5), dpi=100)
                x_axis = np.arange(1, len(data) + 1)
                
                for s, c, a in [(1, 'green', 0.1), (2, 'yellow', 0.15), (3, 'red', 0.1)]:
                    ax.fill_between(x_axis, mean+(s-1)*std_dev, mean+s*std_dev, color=c, alpha=a)
                    ax.fill_between(x_axis, mean-(s-1)*std_dev, mean-s*std_dev, color=c, alpha=a)

                boundaries = [0] + feat['split_locs'] + [len(data)]
                for idx in range(len(boundaries) - 1):
                    s_x, e_x = boundaries[idx], boundaries[idx+1]
                    if idx > 0: ax.axvline(x=s_x + 0.5, color='black', ls='--', lw=1.5, alpha=0.8)
                    if idx % 2 != 0: ax.axvspan(s_x + 0.5, e_x + 0.5, facecolor='#F2F2F2', alpha=0.5, zorder=0)
                
                all_vals = np.concatenate([data, [feat['usl'], feat['lsl'], mean+3*std_dev, mean-3*std_dev]])
                y_min_v, y_max_v = np.min(all_vals), np.max(all_vals)
                y_rng = max(y_max_v - y_min_v, 1e-9)
                y_bottom, y_top = y_min_v - (y_rng * 0.15), y_max_v + (y_rng * 0.15)
                ax.set_ylim(y_bottom, y_top)
                
                ax.axhspan(feat['usl'], y_top, facecolor='none', hatch='////', edgecolor='#FF9999', alpha=0.5)
                ax.axhspan(y_bottom, feat['lsl'], facecolor='none', hatch='////', edgecolor='#FF9999', alpha=0.5)

                ax.set_xlabel('Samples', fontweight='bold')
                ax.set_ylabel('Dimension (in)', fontweight='bold')
                ax.yaxis.set_major_formatter(plt.FormatStrFormatter('%.4f')) 
                
                ax.axhline(mean, color='green')
                ax.axhline(feat['usl'], color='red')
                ax.axhline(feat['lsl'], color='red')
                ax.plot(x_axis, data, marker='o', color='black', lw=1, ms=4)
                
                oot_mask = (data > feat['usl']) | (data < feat['lsl'])
                ax.plot(x_axis[oot_mask], data[oot_mask], 'x', color='red', ms=8, mew=2)
                if is_unstable:
                    ax.plot(x_axis[highlight_indices], data[highlight_indices], 'o', color='red', ms=10, mfc='none', mew=2)

                ax.set_title(f"{name}", fontweight='bold')
                img_path = os.path.join(output_dir, f"TEMP_CHART_{unique_prefix}.png")
                plt.savefig(img_path, bbox_inches='tight')
                plt.close(fig)

                temp_files.append(img_path)
                img = XLImage(img_path)
                
                IMG_WIDTH_PX = 600
                IMG_HEIGHT_PX = 300
                img.width, img.height = IMG_WIDTH_PX, IMG_HEIGHT_PX
                
                ws.add_image(img, f"B{img_start_row}")
                ws.row_dimensions[img_start_row].height = IMG_HEIGHT_PX * 0.75
                
                cell = ws.cell(img_start_row, 1, name)
                cell.font, cell.alignment = Font(bold=True, size=12), Alignment(vertical='top')
                img_start_row += 1

            grey_f, thin_b = PatternFill(start_color="E7E6E6", fill_type="solid"), Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            for scan_row in range(total_insert_count + 1, ws.max_row + 1):
                val_str = str(ws.cell(scan_row, 1).value).strip()
                if val_str in [str(k) for k in metadata_keys if k is not None]:
                    for col_idx in range(1, ws.max_column + 1):
                        c = ws.cell(scan_row, col_idx)
                        c.fill, c.border = grey_f, thin_b
            
            tab_log.append({'name': sheet_name, 'status': 'OK', 'msg': f"Processed {len(features_data)} features"})

        except Exception as e:
            tab_log.append({'name': sheet_name, 'status': 'ERR', 'msg': str(e)})
            continue

    if processed_tabs:
        pdf.page = 1
        pdf.set_y(50)
        pdf.set_font('Arial', 'B', 10)
        pdf.cell(50, 7, "Tab", 1)
        pdf.cell(60, 7, "Feature", 1)
        pdf.cell(30, 7, "Cpk", 1)
        pdf.cell(30, 7, "Status", 1)
        pdf.ln()
        
        pdf.set_font('Arial', '', 10)
        for sheet, name, val, status in pdf_summary_data:
            pdf.cell(50, 7, str(sheet)[:20], 1)
            pdf.cell(60, 7, str(name)[:25], 1)
            pdf.cell(30, 7, val, 1)
            if status == "FAIL": pdf.set_text_color(200, 0, 0)
            else: pdf.set_text_color(0, 128, 0)
            pdf.cell(30, 7, status, 1)
            pdf.set_text_color(0,0,0)
            pdf.ln()

        wb_output.active = 0
        wb_output.save(output_path)
        pdf.output(pdf_path)
        
        for f in temp_files:
            try: os.remove(f)
            except: pass
    
    return {"processed": processed_tabs, "logs": tab_log, "ignored": []}

def main():
    os.system('cls' if os.name == 'nt' else 'clear')
    if os.name == 'nt': os.system(f'title SPC Tool v{TOOL_VERSION}')

    console.print(Panel.fit(
        r"""[bold cyan]SPC GENERATOR[/bold cyan]
[dim]Press Enter to Select One | Use Menu for Multi-Select[/dim]""",
        subtitle=f"v{TOOL_VERSION}"
    ))

    cd = os.getcwd()
    files = [f for f in glob.glob(os.path.join(cd, "SPC-DATA_*.xlsx")) if not os.path.basename(f).startswith("~$")]

    if not files:
        console.print(Panel(
            "[bold red]❌ No input files found![/bold red]\n\n"
            "1. Files must start with: [yellow]SPC-DATA_[/yellow]\n"
            "2. Files must be [yellow].xlsx[/yellow]",
            title="Search Error"
        ))
        questionary.press_any_key_to_continue().ask()
        return

    launcher_options = [
        "[ ▶ PROCESS ALL FILES ]",
        "[ ▶ SELECT MULTIPLE... ]",
        questionary.Separator()
    ] + sorted([os.path.basename(f) for f in files])

    selected_action = questionary.select(
        "Choose Input Action:",
        choices=launcher_options,
        use_indicator=True
    ).ask()

    if not selected_action: return

    files_to_process = []

    if selected_action == "[ ▶ PROCESS ALL FILES ]":
        files_to_process = files
        
    elif selected_action == "[ ▶ SELECT MULTIPLE... ]":
        selected_checkboxes = questionary.checkbox(
            "Select Input Files (Space to select, Enter to confirm):",
            choices=sorted([os.path.basename(f) for f in files]),
            qmark="▶"
        ).ask()
        
        if not selected_checkboxes: return
        files_to_process = [f for f in files if os.path.basename(f) in selected_checkboxes]
        
    else:
        files_to_process = [f for f in files if os.path.basename(f) == selected_action]
    
    if not files_to_process:
        console.print("[red]No files selected. Exiting.[/red]")
        return

    output_root = Path(cd) / "output"
    existing_projects = []
    if output_root.exists():
        existing_projects = sorted([d.name for d in output_root.iterdir() if d.is_dir()])
    
    project_options = existing_projects + ["< Create New Project >"]
    
    selected_option = questionary.select("Select Output Folder:", choices=project_options).ask()
    if selected_option is None: return

    if selected_option == "< Create New Project >":
        project_name = questionary.text("Enter New Project Name:", default="New_Project").ask()
        if project_name is None: return
    else:
        project_name = selected_option
        
    output_dir = output_root / project_name
    output_dir.mkdir(parents=True, exist_ok=True)

    console.print("\n")
    
    for fname in files_to_process:
        filename_only = os.path.basename(fname)
        
        file_tree = Tree(f"[bold cyan]{filename_only}[/bold cyan]")
        
        with console.status(f"[bold green]Processing {filename_only}...[/bold green]"):
            result = process_single_file(fname, output_dir)
        
        if "critical_error" in result:
            file_tree.add(f"[bold red]CRITICAL FAILURE:[/bold red] {result['critical_error']}")
        else:
            logs = result.get('logs', [])
            processed_count = len(result['processed'])
            
            if processed_count == 0:
                file_tree.label = f"[bold red]{filename_only} (NO OUTPUT GENERATED)[/bold red]"
            else:
                file_tree.label = f"[bold green]{filename_only} (Generated)[/bold green]"

            for entry in logs:
                name = entry['name']
                status = entry['status']
                msg = entry['msg']
                
                if status == "OK":
                    file_tree.add(f"[green]✔ {name}[/green]: {msg}")
                elif status == "SKIP":
                    file_tree.add(f"[yellow]⚠ {name} (Skipped)[/yellow]: {msg}")
                elif status == "ERR":
                    file_tree.add(f"[red]❌ {name} (Error)[/red]: {msg}")
        
        console.print(file_tree)
        console.print("") 

    console.print(f"[bold]Output Folder:[/bold] {output_dir}")

if __name__ == "__main__":
    main()