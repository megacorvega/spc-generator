import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# --- CONFIGURATION ---
OUTPUT_FILENAME = "SPC-DATA_Input_Template.xlsx"
METADATA_START_ROW = 2     # Row where metadata fields begin
DATA_TABLE_START_ROW = 8   # Row where the main feature data table begins
NUM_SAMPLE_ROWS = 100      # Number of sample rows in the template

# --- STYLES ---
GREY_FILL = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'), 
    top=Side(style='thin'), bottom=Side(style='thin')
)

def main():
    """
    Generates the Excel input template with styled Nominal/Spec rows.
    """
    
    script_dir = os.getcwd()
    full_path = os.path.join(script_dir, OUTPUT_FILENAME)

    # --- DEFINE TRANSPOSED DATA STRUCTURE ---
    row_labels = ["Nominal", "USL or Upper Tol.", "LSL or Lower Tol."]
    sample_labels = [f"Sample_{i}" for i in range(1, NUM_SAMPLE_ROWS + 1)]
    row_labels.extend(sample_labels)

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "SPC_Data_Input"

        # --- WRITE METADATA HEADER BLOCK ---
        metadata_fields = ["Part Number", "Batch Number", "Date of Inspection", "Inspector", "Notes"]

        ws['A1'] = "SPC ANALYSIS INPUT"
        ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        for i, field in enumerate(metadata_fields):
            row_num = METADATA_START_ROW + i
            ws.cell(row=row_num, column=1, value=field).font = Font(bold=True)
            ws.cell(row=row_num, column=2, value="[INPUT HERE]")

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20

        # --- WRITE TRANSPOSED DATA TABLE ---
        
        # 1. Header Row
        header_cell = ws.cell(row=DATA_TABLE_START_ROW, column=1, value="Feature Name")
        header_cell.font = Font(bold=True, color="FFFFFF")
        header_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_cell.alignment = Alignment(horizontal='center')

        # Add feature columns
        example_features = [f"Dim {i}" for i in range(1, 11)]
        for col_idx, feature_name in enumerate(example_features, 2):
            cell = ws.cell(row=DATA_TABLE_START_ROW, column=col_idx, value=feature_name)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')

        # 2. Rows (Specs + Samples)
        for row_offset, label in enumerate(row_labels):
            row_num = DATA_TABLE_START_ROW + 1 + row_offset
            
            # Check if this is a spec row that needs styling
            is_spec_row = label in ["Nominal", "USL", "LSL"]
            
            # Label Column
            label_cell = ws.cell(row=row_num, column=1, value=label)
            label_cell.font = Font(bold=True)
            label_cell.alignment = Alignment(horizontal='left')

            if is_spec_row:
                label_cell.fill = GREY_FILL
                label_cell.border = THIN_BORDER
                # specific text colors
                if label == "Nominal": label_cell.font = Font(bold=True, color="000000") # Black
                if label in ["USL", "LSL"]: label_cell.font = Font(bold=True, color="C0504D") # Red

            # Empty Data Cells
            for col_idx in range(2, 2 + len(example_features)):
                cell = ws.cell(row=row_num, column=col_idx, value="")
                cell.alignment = Alignment(horizontal='center')
                
                # Apply styling to input cells for specs
                if is_spec_row:
                    cell.fill = GREY_FILL
                    cell.border = THIN_BORDER

        wb.save(full_path)
        print("-" * 40)
        print(f"SUCCESS: Template created at: {OUTPUT_FILENAME}")
        print("-" * 40)
        
    except PermissionError:
        print(f"ERROR: Close '{OUTPUT_FILENAME}' and try again.")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")

if __name__ == "__main__":
    main()