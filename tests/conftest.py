"""
Pytest configuration and shared fixtures for SPC Generator tests.

This file is automatically loaded by pytest and provides project-wide
test configuration and fixtures.
"""
import pytest
import warnings
import matplotlib
import os
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
import tempfile
import shutil
from datetime import datetime


def pytest_configure(config):
    """
    Pytest configuration hook - runs before test collection.
    """
    # Use non-interactive matplotlib backend for testing
    matplotlib.use('Agg')

    # Suppress matplotlib warnings during tests
    warnings.filterwarnings('ignore', category=UserWarning, module='matplotlib')

    # Register custom markers
    config.addinivalue_line(
        "markers", "slow: marks tests as slow (deselect with '-m \"not slow\"')"
    )
    config.addinivalue_line(
        "markers", "integration: marks tests as integration tests"
    )
    config.addinivalue_line(
        "markers", "unit: marks tests as unit tests"
    )


@pytest.fixture(autouse=True)
def reset_matplotlib():
    """Automatically reset matplotlib state between tests."""
    import matplotlib.pyplot as plt
    yield
    plt.close('all')


@pytest.fixture(autouse=True)
def suppress_console_output(monkeypatch):
    """
    Suppress print statements during tests (optional).
    Remove autouse=True if you want to see print output during testing.
    """
    # You can uncomment these lines to suppress output
    # import sys
    # from io import StringIO
    # monkeypatch.setattr(sys, 'stdout', StringIO())
    pass


@pytest.fixture
def mock_input(monkeypatch):
    """
    Factory fixture to mock input() calls.

    Usage:
        def test_something(mock_input):
            mock_input(['yes', 'no'])
            # Your test code that calls input()
    """
    def _mock_input(inputs):
        inputs_iter = iter(inputs)
        monkeypatch.setattr('builtins.input', lambda _: next(inputs_iter))
    return _mock_input


@pytest.fixture
def no_sleep(monkeypatch):
    """Mock time.sleep to speed up tests that have delays."""
    import time
    monkeypatch.setattr(time, 'sleep', lambda x: None)


# Environment information
@pytest.fixture(scope="session")
def test_environment():
    """Provides test environment information."""
    return {
        'python_version': os.sys.version,
        'platform': os.sys.platform,
        'cwd': os.getcwd()
    }


# ============================================================================
# SHARED DATA FIXTURES
# ============================================================================

@pytest.fixture
def sample_metadata():
    """Provides sample metadata for SPC reports."""
    return {
        "Part Number": "PN-12345",
        "Batch Number": "BATCH-001",
        "Date of Inspection": datetime(2025, 1, 15),
        "Inspector": "John Doe"
    }


@pytest.fixture
def sample_feature_data():
    """Provides a list of feature data dictionaries for testing."""
    np.random.seed(42)  # For reproducible tests
    return [
        {
            'name': 'Diameter',
            'nominal': 10.0,
            'usl': 10.5,
            'lsl': 9.5,
            'subgroup': 5,
            'data': np.random.normal(10.0, 0.1, 50)
        },
        {
            'name': 'Length',
            'nominal': 25.0,
            'usl': 25.3,
            'lsl': 24.7,
            'subgroup': 3,
            'data': np.random.normal(25.0, 0.05, 30)
        },
        {
            'name': 'Weight',
            'nominal': 100.0,
            'usl': 105.0,
            'lsl': 95.0,
            'subgroup': 10,
            'data': np.random.normal(100.0, 1.5, 100)
        }
    ]


@pytest.fixture
def minimal_feature_data():
    """Feature with minimal data (edge case testing)."""
    return {
        'name': 'Minimal_Feature',
        'nominal': 5.0,
        'usl': 5.5,
        'lsl': 4.5,
        'subgroup': 3,
        'data': np.array([5.0])  # Only 1 data point
    }


@pytest.fixture
def insufficient_subgroup_data():
    """Feature with insufficient subgroups for within-variation calculation."""
    return {
        'name': 'Insufficient_Subgroups',
        'nominal': 15.0,
        'usl': 16.0,
        'lsl': 14.0,
        'subgroup': 5,
        'data': np.array([15.1, 15.0, 14.9])  # < 2 subgroups
    }


@pytest.fixture
def temp_dir():
    """Creates a temporary directory for test files."""
    temp_path = tempfile.mkdtemp(prefix="spc_test_")
    yield temp_path
    # Cleanup after tests
    shutil.rmtree(temp_path, ignore_errors=True)


@pytest.fixture
def create_spc_excel_file(temp_dir, sample_metadata):
    """Factory fixture to create properly formatted SPC Excel input files."""
    def _create_file(filename="SPC-DATA_test.xlsx", features=None):
        """
        Creates a test Excel file in SPC-DATA format.

        Args:
            filename: Name of the Excel file to create
            features: List of dicts with keys: name, nominal, usl, lsl, subgroup, samples

        Returns:
            Full path to created Excel file
        """
        filepath = os.path.join(temp_dir, filename)
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        # Header Section (rows 1-6)
        ws['A1'] = "SPC Data Input"
        ws['A1'].font = Font(bold=True, size=14)

        ws['A2'] = "Part Number:"
        ws['B2'] = sample_metadata["Part Number"]
        ws['A3'] = "Batch Number:"
        ws['B3'] = sample_metadata["Batch Number"]
        ws['A4'] = "Date of Inspection:"
        ws['B4'] = sample_metadata["Date of Inspection"]
        ws['A5'] = "Inspector:"
        ws['B5'] = sample_metadata["Inspector"]

        # Data section starts at row 7
        ws['A7'] = "Feature Name"

        if features is None:
            # Default: Create 2 simple features
            features = [
                {
                    'name': 'Test_Feature_1',
                    'nominal': 10.0,
                    'usl': 10.5,
                    'lsl': 9.5,
                    'subgroup': 5,
                    'samples': [10.1, 10.0, 9.9, 10.2, 9.8, 10.0, 10.1, 9.9, 10.0, 10.1]
                },
                {
                    'name': 'Test_Feature_2',
                    'nominal': 25.0,
                    'usl': 26.0,
                    'lsl': 24.0,
                    'subgroup': 3,
                    'samples': [25.1, 25.0, 24.9, 25.2, 24.8, 25.0]
                }
            ]

        # Write feature columns
        for col_idx, feature in enumerate(features, start=2):  # Start from column B
            col_letter = chr(65 + col_idx - 1)  # B, C, D, etc.

            ws[f'{col_letter}7'] = feature['name']
            ws[f'{col_letter}8'] = feature['nominal']
            ws[f'{col_letter}9'] = feature['usl']
            ws[f'{col_letter}10'] = feature['lsl']
            ws[f'{col_letter}11'] = feature['subgroup']

            # Write sample data starting from row 12
            for sample_idx, value in enumerate(feature['samples'], start=12):
                ws[f'{col_letter}{sample_idx}'] = value

        # Write row labels
        ws['A8'] = 'Nominal'
        ws['A9'] = 'USL'
        ws['A10'] = 'LSL'
        ws['A11'] = 'Subgroup Size'

        # Write sample row labels
        for i, sample_num in enumerate(range(1, 51), start=12):
            ws[f'A{i}'] = f'Sample_{sample_num}'

        wb.save(filepath)
        return filepath

    return _create_file


@pytest.fixture
def normal_distribution_data():
    """Generates normally distributed data for testing."""
    np.random.seed(123)
    return np.random.normal(loc=10.0, scale=0.5, size=100)


@pytest.fixture
def subgroup_test_data():
    """Provides data specifically designed for subgroup testing."""
    # Create 10 subgroups of size 5
    np.random.seed(456)
    subgroups = []
    for i in range(10):
        subgroup = np.random.normal(loc=10.0 + i*0.1, scale=0.2, size=5)
        subgroups.extend(subgroup)
    return np.array(subgroups)


@pytest.fixture
def edge_case_data():
    """Provides various edge case datasets."""
    return {
        'empty': np.array([]),
        'single': np.array([5.0]),
        'constant': np.array([10.0] * 20),
        'two_points': np.array([9.5, 10.5]),
        'with_nan': np.array([10.0, np.nan, 10.5, 9.5]),
        'outliers': np.array([10.0, 10.1, 10.0, 50.0, 10.0, -20.0, 10.1])
    }


@pytest.fixture
def c4_constant_tests():
    """Provides test cases for c4 constant validation."""
    return {
        2: 0.7979,
        5: 0.9400,
        10: 0.9727,
        15: 0.9823,
        16: None,  # Out of range
        1: None,   # Out of range
    }


@pytest.fixture
def filename_sanitization_cases():
    """Test cases for filename sanitization."""
    return [
        ('Feature: Test/Name', 'Feature Test Name'),
        ('Test*Feature?', 'TestFeature'),
        ('Normal_Name', 'Normal_Name'),
        ('Feature<>|"Test', 'FeatureTest'),
        ('Path\\With\\Slashes', 'PathWithSlashes'),
    ]


@pytest.fixture
def duplicate_file_setup(temp_dir):
    """Creates files for testing unique filepath generation."""
    base_path = os.path.join(temp_dir, "test_file.xlsx")

    # Create initial file and duplicates
    open(base_path, 'w').close()
    open(os.path.join(temp_dir, "test_file_1.xlsx"), 'w').close()
    open(os.path.join(temp_dir, "test_file_2.xlsx"), 'w').close()

    return {
        'directory': temp_dir,
        'base_name': 'test_file.xlsx',
        'expected_next': os.path.join(temp_dir, "test_file_3.xlsx")
    }
